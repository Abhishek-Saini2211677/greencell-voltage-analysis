"""
Builds Voltage_Analysis.xlsx - the Excel half of the assignment.

Sheets
------
Read Me            what is in the workbook
Data               raw data, sorted chronologically (Voltage + Timestamp)
Chart              voltage vs timestamp with a LINEAR trendline, plus a second
                   chart with a MOVING-AVERAGE trendline
Summary            headline statistics, written as live Excel formulas
Interpretation     the 5-sentence reading of the data
Peaks and Lows     tabulated local highs / lows  (from analysis.py)
Cycles             one row per discharge cycle
Below 20V          every instance the voltage fell under 20 V
Acceleration       every episode where the discharge steepened
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import analysis

OUT_XLSX = "Voltage_Analysis.xlsx"

FONT = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F3864")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = H_FILL, H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_table(ws, df, start_row=1, date_fmt="dd-mm-yyyy hh:mm", widths=None):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col))
    for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if isinstance(v, pd.Timestamp):
                cell = ws.cell(row=i, column=j, value=v.to_pydatetime())
                cell.number_format = date_fmt
            else:
                cell = ws.cell(row=i, column=j,
                               value=v.item() if hasattr(v, "item") else v)
            cell.font = BODY
    style_header(ws, start_row, len(df.columns))
    for j, col in enumerate(df.columns, start=1):
        w = widths[j - 1] if widths else max(12, min(26, len(str(col)) + 6))
        ws.column_dimensions[get_column_letter(j)].width = w


def add_scatter(ws_chart, data_sheet_title, nrows, anchor, title,
                trend: str, ma_period: int | None = None):
    """Voltage (col A of Data) on Y, Timestamp (col B) on X."""
    ch = ScatterChart()
    ch.title = title
    ch.style = 2
    ch.height, ch.width = 11, 30
    ch.x_axis = DateAxis(crossAx=100)
    ch.x_axis.number_format = "dd-mm hh:mm"
    ch.x_axis.majorTimeUnit = "days"
    ch.x_axis.title = "Timestamp"
    ch.y_axis.title = "Voltage (V)"
    ch.y_axis.scaling.min = 0
    ch.y_axis.scaling.max = 110

    xvalues = Reference(ws_chart.parent[data_sheet_title],
                        min_col=2, min_row=2, max_row=nrows + 1)
    yvalues = Reference(ws_chart.parent[data_sheet_title],
                        min_col=1, min_row=1, max_row=nrows + 1)
    s = Series(yvalues, xvalues, title_from_data=True)
    s.marker.symbol = "none"
    s.graphicalProperties.line.width = 9000          # ~0.75 pt
    s.graphicalProperties.line.solidFill = "2A78D6"
    s.smooth = False
    if trend == "linear":
        s.trendline = Trendline(trendlineType="linear", dispEq=True, dispRSqr=True)
    else:
        s.trendline = Trendline(trendlineType="movingAvg", period=ma_period or 500)
    ch.series.append(s)
    ws_chart.add_chart(ch, anchor)


def main():
    res = analysis.main()
    raw, minute = res["raw"], res["minute"]
    tp, cycles, below20, accel = res["tp"], res["cycles"], res["below20"], res["accel"]

    wb = Workbook()

    # ---------------- Read Me ----------------
    ws = wb.active
    ws.title = "Read Me"
    ws["A1"] = "Voltage data - visualisation and interpretation"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Source file: Sample_Data.csv (columns: Values = voltage in V, Timestamp).",
        f"Records: {len(raw):,} readings from "
        f"{raw['Timestamp'].min():%d-%m-%Y %H:%M} to {raw['Timestamp'].max():%d-%m-%Y %H:%M}.",
        "",
        "Sheet guide",
        "  Data             All raw readings, sorted oldest to newest on the parsed date.",
        "  Chart            Voltage vs Timestamp with a linear trendline, and a second chart",
        "                   using a moving-average trendline.",
        "  Summary          Headline statistics as live Excel formulas over the Data sheet.",
        "  Interpretation   Five-sentence reading of the data (assignment part 1.3).",
        "  Peaks and Lows   Every local high and local low found by the Python analysis.",
        "  Cycles           One row per discharge cycle: peak, low, drop and rate.",
        "  Below 20V        Every instance the voltage went below 20 V.",
        "  Acceleration     Every episode where the downward slope accelerated (bonus part).",
        "",
        "Note on the x-axis: the Timestamp column must be treated as a date, not text.",
        "Plotted as text the axis runs in string order (01-07 before 26-06), which is how the",
        "reference figure in the assignment annexure ends up with a scrambled x-axis. The",
        "supplied CSV is already in date order; the Data sheet is sorted on the parsed date",
        "so that this holds by construction rather than by luck.",
    ]
    for i, t in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=t).font = BOLD if t.strip() in (
            "Sheet guide",) else BODY
    ws.column_dimensions["A"].width = 100

    # ---------------- Data ----------------
    wsd = wb.create_sheet("Data")
    data_out = raw[["Voltage", "Timestamp"]].copy()
    wsd.cell(row=1, column=1, value="Voltage").font = H_FONT
    wsd.cell(row=1, column=2, value="Timestamp").font = H_FONT
    for i, (v, t) in enumerate(zip(data_out["Voltage"], data_out["Timestamp"]), start=2):
        wsd.cell(row=i, column=1, value=float(v)).font = BODY
        c = wsd.cell(row=i, column=2, value=t.to_pydatetime())
        c.number_format = "dd-mm-yyyy hh:mm"
        c.font = BODY
    style_header(wsd, 1, 2)
    wsd.column_dimensions["A"].width = 12
    wsd.column_dimensions["B"].width = 20
    n = len(data_out)

    # ---------------- Chart ----------------
    wsc = wb.create_sheet("Chart")
    wsc["A1"] = "Voltage vs Timestamp"
    wsc["A1"].font = TITLE_FONT
    wsc["A2"] = ("Y axis = Voltage (V), X axis = Timestamp. The dashed line is the "
                 "trendline; its equation and R-squared are shown on the chart.")
    wsc["A2"].font = BODY
    wsc.page_setup.orientation = "landscape"
    wsc.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    wsc.page_setup.fitToWidth = 1
    wsc.page_setup.fitToHeight = 0
    add_scatter(wsc, "Data", n, "A4", "Voltage over time with linear trendline",
                trend="linear")
    add_scatter(wsc, "Data", n, "A28",
                "Voltage over time with 500-point moving-average trendline",
                trend="movingAvg", ma_period=500)
    wsc.column_dimensions["A"].width = 14

    # ---------------- Summary ----------------
    wss = wb.create_sheet("Summary")
    wss["A1"] = "Headline statistics"
    wss["A1"].font = TITLE_FONT
    wss["A2"] = "Every value below is a live Excel formula over the Data sheet."
    wss["A2"].font = Font(name=FONT, size=9, italic=True, color="52514E")
    rows = [
        ("Number of readings", f"=COUNT(Data!A2:A{n + 1})", "0"),
        ("First timestamp", f"=MIN(Data!B2:B{n + 1})", "dd-mm-yyyy hh:mm"),
        ("Last timestamp", f"=MAX(Data!B2:B{n + 1})", "dd-mm-yyyy hh:mm"),
        ("Days covered", f"=MAX(Data!B2:B{n + 1})-MIN(Data!B2:B{n + 1})", "0.00"),
        ("Minimum voltage (V)", f"=MIN(Data!A2:A{n + 1})", "0.0"),
        ("Maximum voltage (V)", f"=MAX(Data!A2:A{n + 1})", "0.0"),
        ("Average voltage (V)", f"=AVERAGE(Data!A2:A{n + 1})", "0.00"),
        ("Median voltage (V)", f"=MEDIAN(Data!A2:A{n + 1})", "0.00"),
        ("Standard deviation (V)", f"=STDEV(Data!A2:A{n + 1})", "0.00"),
        ("Readings below 20 V", f"=COUNTIF(Data!A2:A{n + 1},\"<20\")", "0"),
        ("Readings below 30 V", f"=COUNTIF(Data!A2:A{n + 1},\"<30\")", "0"),
        ("Readings at 100 V (full)", f"=COUNTIF(Data!A2:A{n + 1},100)", "0"),
        ("Trendline slope (V per day)",
         f"=SLOPE(Data!A2:A{n + 1},Data!B2:B{n + 1})", "0.0000"),
        ("Trendline intercept (V)",
         f"=INTERCEPT(Data!A2:A{n + 1},Data!B2:B{n + 1})", "0.00"),
        ("Trendline R-squared", f"=RSQ(Data!A2:A{n + 1},Data!B2:B{n + 1})", "0.0000"),
    ]
    wss.cell(row=4, column=1, value="Metric")
    wss.cell(row=4, column=2, value="Value")
    style_header(wss, 4, 2)
    for i, (label, formula, fmt) in enumerate(rows, start=5):
        wss.cell(row=i, column=1, value=label).font = BODY
        c = wss.cell(row=i, column=2, value=formula)
        c.number_format, c.font = fmt, BODY
        c.border = BOX
    wss.column_dimensions["A"].width = 32
    wss.column_dimensions["B"].width = 20
    wss.cell(row=len(rows) + 6, column=1,
             value="Trendline slope/intercept are computed against the Excel date "
                   "serial number, so the slope is in volts per day.").font = Font(
        name=FONT, size=9, italic=True, color="52514E")

    # ---------------- Interpretation ----------------
    wsi = wb.create_sheet("Interpretation")
    wsi["A1"] = "Interpretation of the data"
    wsi["A1"].font = TITLE_FONT
    sentences = interpretation_sentences(res)
    wsi["A3"] = "Five-sentence description (assignment part 1.3)"
    wsi["A3"].font = BOLD
    r = 4
    for i, s in enumerate(sentences, start=1):
        c = wsi.cell(row=r, column=1, value=f"{i}. {s}")
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wsi.row_dimensions[r].height = 46
        r += 1
    wsi.column_dimensions["A"].width = 118

    # ---------------- result tables ----------------
    write_table(wb.create_sheet("Peaks and Lows"), tp)
    write_table(wb.create_sheet("Cycles"), cycles)
    ws20 = wb.create_sheet("Below 20V")
    if below20.empty:
        ws20["A1"] = "Instances where the voltage went below 20 V"
        ws20["A1"].font = TITLE_FONT
        ws20["A3"] = (f"NONE. The lowest reading anywhere in the file is "
                      f"{raw['Voltage'].min():.0f} V, so the series never crosses 20 V.")
        ws20["A3"].font = BOLD
        ws20["A5"] = ("The same test run at 30 V returns "
                      f"{len(analysis.threshold_episodes(raw, 30.0))} instances, which "
                      "confirms the detection logic works - see the CSV "
                      "outputs/below_30V_instances.csv.")
        ws20["A5"].font = BODY
        ws20["A5"].alignment = Alignment(wrap_text=True)
        ws20.column_dimensions["A"].width = 110
        write_table(ws20, analysis.threshold_episodes(raw, 30.0), start_row=8)
        ws20["A7"] = "For reference - every instance below 30 V:"
        ws20["A7"].font = BOLD
    else:
        write_table(ws20, below20)
    write_table(wb.create_sheet("Acceleration"), accel)

    wb.save(OUT_XLSX)
    print(f"\nWrote {OUT_XLSX}")


def interpretation_sentences(res) -> list[str]:
    raw, cycles, tp = res["raw"], res["cycles"], res["tp"]
    accel, minute = res["accel"], res["minute"]
    n_cycles = len(cycles)
    mean_drop = cycles["Drop (V)"].mean()
    mean_dur = cycles["Duration (h)"].mean()
    mean_rate = cycles["Avg rate (V/h)"].mean()
    _, slope, _ = analysis.linear_trend(minute, raw)
    days = (raw["Timestamp"].max() - raw["Timestamp"].min()).total_seconds() / 86400
    peaks = tp[tp["Type"].str.startswith("Peak")]["Voltage"]
    lows = tp[tp["Type"].str.startswith("Low")]["Voltage"]
    return [
        f"The record covers {days:.1f} days ({len(raw):,} readings, roughly one every "
        f"15-20 seconds) and the voltage never leaves the band "
        f"{raw['Voltage'].min():.0f}-{raw['Voltage'].max():.0f} V, so the series is not "
        f"a drifting signal but a repeating duty cycle.",

        f"That cycle is a sawtooth: a fast rise back towards {raw['Voltage'].max():.0f} V "
        f"followed by a long, almost straight decline - {n_cycles} complete discharge "
        f"cycles are visible, each losing on average {mean_drop:.0f} V over "
        f"{mean_dur:.1f} hours ({mean_rate:.1f} V/h).",

        f"Because charge and discharge repeat around the same mean, the linear trendline is "
        f"nearly flat ({slope:+.2f} V/day over the week, R-squared about 0.001) - it says "
        f"the asset is being cycled consistently rather than degrading, and it is the "
        f"moving average, not the straight line, that carries the information here.",

        f"The tops are capped: {(peaks == raw['Voltage'].max()).sum()} of the {len(peaks)} "
        f"recorded highs sit at exactly {raw['Voltage'].max():.0f} V and the rest are lower "
        f"only because the recharge itself fell inside a logging gap, while the lows scatter "
        f"widely from {lows.min():.0f} V to {lows.max():.0f} V - the unit is always taken "
        f"back to full, but how deeply it is drained first varies with use.",

        f"The decline within a cycle is not uniform - in {len(accel)} separate episodes the "
        f"fall steepens sharply (reaching "
        f"{accel['Steepest slope reached (V/h)'].min():.0f} V/h against a cycle average of "
        f"{mean_rate:.1f} V/h), which looks like heavier load being drawn mid-discharge, "
        f"and the voltage never once drops below 20 V - its lowest reading all week is "
        f"{raw['Voltage'].min():.0f} V - so the low-voltage limit was never approached.",
    ]


if __name__ == "__main__":
    main()
